#!/usr/bin/env python3
"""Build a readable analysis workbook for the native-CR sensitivity results."""

from __future__ import annotations

import csv
import json
import math
import statistics
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
WORKSPACE = REPO.parent
RESULTS = HERE / "results"
FIGURES = HERE / "figures"
OUTPUT = RESULTS / "native_cr_no_btv_analysis_20260820.xlsx"

NAVY = "1F4E78"
BLUE = "D9EAF7"
PALE_BLUE = "EDF4FA"
GREEN = "E2F0D9"
GREEN_STRONG = "C6E0B4"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
TEXT = "1F2937"
GRID = "B7C9D6"
BODY_FONT = Font(name="Arial", size=12, color=TEXT)
HEADER_FONT = Font(name="Arial", size=13, bold=True, color=WHITE)
TITLE_FONT = Font(name="Arial", size=18, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Arial", size=12, italic=True, color=TEXT)
THIN = Side(style="thin", color=GRID)

EARTH_ORDER = ("month", "qflag", "decade", "temp_bin", "year")
JOB_ORDER = (
    "mc_ctype", "kind", "role", "pi_itype", "mi_itype", "year", "country",
    "movie_keyword_movie_id", "cast_info_movie_id", "cast_info_person_id",
)
CLUSTER_ORDER = (
    "iid", "f2", "f4", "f8", "f16", "f32", "f64", "f128", "f256",
    "f1024", "sorted",
)
OPERATIONS = ("OR", "AND", "NOT", "COMP")
BACKENDS = ("DDC", "CRoaring", "WAH", "EWAH")
DENSITIES = ("0.1%", "0.5%", "1%", "2%", "5%", "10%", "20%", "30%", "40%", "50%")
DENSITY_BACKENDS = ("DDC", "CRoaring", "WAH", "EWAH", "Bitset-AVX512", "Concise")
PLOTTED_CLUSTER = {"iid", "f4", "f16", "f64", "f256"}
RAW_INTEGER_FIELDS = {"num_rows", "replicates"}
RAW_FLOAT_FIELDS = {
    "time_ms", "latency_over_ddc", "min_ms", "max_ms", "cf1",
    "throughput_ops_s", "or_time_ms", "native_cr_ms",
    "same_run_delivered_cr_ms", "historical_delivered_cr_ms",
    "native_over_same_run_delivered", "native_speedup_from_removing_delivery",
    "delivery_overhead_ms", "delivery_share_pct",
    "same_run_delivered_over_historical",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def geometric_mean(values: list[float]) -> float:
    return math.exp(statistics.mean(math.log(value) for value in values))


def title(ws, text: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1, text)
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 42
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 105


def section(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Arial", size=14, bold=True, color=TEXT)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 26


def write_table(
    ws,
    start_row: int,
    headers: list[str],
    rows: list[list[object]],
    widths: dict[int, float] | None = None,
    freeze: str | None = None,
) -> tuple[int, dict[str, int]]:
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(start_row, column, value)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[start_row].height = 38
    for row_offset, values in enumerate(rows, start=1):
        row_number = start_row + row_offset
        ws.row_dimensions[row_number].height = 24
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9E2F3"))
            if row_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
    end_row = start_row + len(rows)
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{end_row}"
    ws.freeze_panes = freeze or f"A{start_row + 1}"
    header_map = {name: index + 1 for index, name in enumerate(headers)}
    if widths:
        for column, width in widths.items():
            ws.column_dimensions[get_column_letter(column)].width = width
    else:
        for column, header in enumerate(headers, start=1):
            maximum = len(header)
            for values in rows[:300]:
                maximum = max(maximum, len(str(values[column - 1] if column - 1 < len(values) else "")))
            ws.column_dimensions[get_column_letter(column)].width = min(max(maximum + 2, 12), 46)
    return end_row, header_map


def ratio_fill(cell, value: float) -> None:
    if value < 0.95:
        color = GREEN
    elif value <= 1.05:
        color = PALE_BLUE
    elif value <= 2.0:
        color = YELLOW
    else:
        color = RED
    cell.fill = PatternFill("solid", fgColor=color)


def apply_number_format(ws, rows: range, columns: list[int], fmt: str) -> None:
    for row in rows:
        for column in columns:
            ws.cell(row, column).number_format = fmt


def build_overview(wb: Workbook, data: dict[str, object]) -> None:
    ws = wb.create_sheet("00 Overview")
    title(
        ws,
        "Native CRoaring / No-Dense-Delivery Analysis",
        "Readable analysis workbook. Fresh measurements apply to CRoaring only; DDC, WAH, and EWAH are frozen baselines unless explicitly stated.",
        8,
    )
    section(ws, 4, "Key results", 8)
    headers = ["Metric", "New native result", "Previous delivered result", "Change / interpretation"]
    metrics = data["overview_metrics"]
    rows = [[item[0], item[1], item[2], item[3]] for item in metrics]
    end, _ = write_table(ws, 5, headers, rows, {1: 39, 2: 23, 3: 25, 4: 62}, "A6")
    for row in range(6, end + 1):
        ws.row_dimensions[row].height = 40
        ws.cell(row, 2).font = Font(name="Arial", size=12, bold=True, color=TEXT)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=GREEN)

    note_row = end + 2
    section(ws, note_row, "Measurement contract and caveats", 8)
    notes = [
        "CRoaring OR/AND/NOT use native *_op rows; COMP uses the newly measured COMP_op_native row. No CR-to-dense-BTV delivery is included.",
        "Earth, JOB, and Cluster: CRoaring is a fresh 3-process median. DDC/WAH/EWAH are frozen baselines and retain their own result representations.",
        "Density: 55 upper-triangle CR points were freshly measured and mirrored into a 100-cell matrix. Other backends use the frozen density grid.",
        "JOB COMP remains a sensitivity comparison because the frozen DDC baseline used a shorter equivalent execution plan.",
        "Latency is milliseconds. Six decimal places are retained so sub-microsecond JOB values are never displayed as zero.",
    ]
    for offset, text in enumerate(notes, start=1):
        cell = ws.cell(note_row + offset, 1, f"• {text}")
        ws.merge_cells(start_row=note_row + offset, start_column=1, end_row=note_row + offset, end_column=8)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.fill = PatternFill("solid", fgColor=YELLOW if offset in (3, 4) else PALE_BLUE)
        ws.row_dimensions[note_row + offset].height = 32

    chart_row = note_row + len(notes) + 2
    section(ws, chart_row, "Table 4 total payload comparison (MiB)", 8)
    ws.cell(chart_row + 1, 1, "DDC")
    ws.cell(chart_row + 2, 1, "CR plain")
    ws.cell(chart_row + 3, 1, "CR run")
    ws.cell(chart_row + 1, 2, data["table4_totals"]["ddc"])
    ws.cell(chart_row + 2, 2, data["table4_totals"]["plain"])
    ws.cell(chart_row + 3, 2, data["table4_totals"]["run"])
    for row in range(chart_row + 1, chart_row + 4):
        ws.cell(row, 1).font = BODY_FONT
        ws.cell(row, 2).font = BODY_FONT
        ws.cell(row, 2).number_format = "0.000"
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Total native payload"
    chart.y_axis.title = "MiB"
    chart.add_data(Reference(ws, min_col=2, min_row=chart_row + 1, max_row=chart_row + 3), titles_from_data=False)
    chart.set_categories(Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + 3))
    chart.height = 7.2
    chart.width = 13.5
    chart.legend = None
    ws.add_chart(chart, f"D{chart_row + 1}")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 23


def build_table4(wb: Workbook, data: dict[str, object]) -> None:
    ws = wb.create_sheet("01 Table4 Run Size")
    title(
        ws,
        "Table 4 Analysis: Run-Optimized CRoaring Payload",
        "MiB uses 2^20 bytes. CR plain/run are fresh measurements; DDC/WAH/EWAH are frozen payload counters. Winner compares all four backends.",
        17,
    )
    headers = [
        "Field", "N", "Card.", "Avg. density", "CF1", "Gini",
        "DDC MiB", "WAH MiB", "EWAH MiB", "CR plain MiB", "CR run MiB",
        "CR run bytes", "Run reduction", "CR/DDC ratio", "CR Inc. vs DDC",
        "Overall winner", "Measurement status",
    ]
    rows = data["table4_rows"]
    end, h = write_table(
        ws, 4, headers, rows,
        {1: 42, 2: 15, 3: 13, 4: 15, 5: 14, 6: 12, 7: 14, 8: 14,
         9: 14, 10: 15, 11: 15, 12: 18, 13: 16, 14: 15, 15: 17,
         16: 16, 17: 34},
        "G5",
    )
    apply_number_format(ws, range(5, end + 1), [2, 3, 12], "#,##0")
    apply_number_format(ws, range(5, end + 1), [4], "0.000000%")
    apply_number_format(ws, range(5, end + 1), [5], "#,##0.00")
    apply_number_format(ws, range(5, end + 1), [6], "0.000")
    apply_number_format(ws, range(5, end + 1), [7, 8, 9, 10, 11], "0.000")
    apply_number_format(ws, range(5, end + 1), [13, 15], "0.0%")
    apply_number_format(ws, range(5, end + 1), [14], "0.000x")
    for row in range(5, end + 1):
        ws.row_dimensions[row].height = 34
        ratio_fill(ws.cell(row, h["CR/DDC ratio"]), float(ws.cell(row, h["CR/DDC ratio"]).value))
        ws.cell(row, h["Overall winner"]).fill = PatternFill("solid", fgColor=GREEN_STRONG)
        ws.cell(row, h["Overall winner"]).font = Font(name="Arial", size=12, bold=True)
    ws.conditional_formatting.add(
        f"M5:M{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"),
    )


def logical_analysis_sheet(wb: Workbook, name: str, rows: list[dict[str, str]], order: tuple[str, ...]) -> None:
    ws = wb.create_sheet(name)
    title(
        ws,
        name.replace("_", " "),
        "Latency in milliseconds. CRoaring is freshly measured native output (3 process replicates); DDC/WAH/EWAH are frozen baselines.",
        18,
    )
    by = {(row["case"], row["operation"], row["backend"]): row for row in rows}
    anchors = data_global["anchor_map"]
    output = []
    for case in order:
        for operation in OPERATIONS:
            backend_rows = [by[(case, operation, backend)] for backend in BACKENDS]
            cr = by[(case, operation, "CRoaring")]
            anchor = anchors.get((rows[0]["group"], case, operation), {})
            output.append([
                case, int(cr["num_rows"]), operation,
                *[float(item["time_ms"]) for item in backend_rows],
                *[float(by[(case, operation, backend)]["latency_over_ddc"]) for backend in ("CRoaring", "WAH", "EWAH")],
                float(anchor["same_run_delivered_cr_ms"]) if anchor else None,
                float(anchor["delivery_share_pct"]) / 100.0 if anchor else None,
                cr["winner"], int(cr["replicates"]), float(cr["min_ms"]), float(cr["max_ms"]),
                cr["source_operation"], "Fresh CR / frozen non-CR baselines",
            ])
    headers = [
        "Case", "N", "Operation", "DDC ms", "CR native ms", "WAH ms", "EWAH ms",
        "CR/DDC", "WAH/DDC", "EWAH/DDC", "CR delivered ms", "Delivery share",
        "Winner", "CR reps", "CR min ms", "CR max ms", "CR source operation", "Measurement status",
    ]
    end, h = write_table(
        ws, 4, headers, output,
        {1: 27, 2: 15, 3: 13, 4: 15, 5: 16, 6: 15, 7: 15, 8: 13,
         9: 13, 10: 13, 11: 17, 12: 16, 13: 15, 14: 12, 15: 15,
         16: 15, 17: 20, 18: 35},
        "D5",
    )
    apply_number_format(ws, range(5, end + 1), [2], "#,##0")
    apply_number_format(ws, range(5, end + 1), [4, 5, 6, 7, 11, 15, 16], "0.000000")
    apply_number_format(ws, range(5, end + 1), [8, 9, 10], "0.000x")
    apply_number_format(ws, range(5, end + 1), [12], "0.0%")
    for row in range(5, end + 1):
        ws.row_dimensions[row].height = 30
        for column in (8, 9, 10):
            ratio_fill(ws.cell(row, column), float(ws.cell(row, column).value))
        winner = ws.cell(row, h["Winner"]).value
        time_column = {"DDC": 4, "CRoaring": 5, "WAH": 6, "EWAH": 7}[winner]
        ws.cell(row, time_column).fill = PatternFill("solid", fgColor=GREEN_STRONG)
        ws.cell(row, time_column).font = Font(name="Arial", size=12, bold=True)


def build_cluster(wb: Workbook, rows: list[dict[str, str]], throughput: bool) -> None:
    name = "05 Cluster Throughput" if throughput else "04 Cluster Latency"
    ws = wb.create_sheet(name)
    unit = "operations per second" if throughput else "milliseconds"
    title(
        ws,
        name,
        f"Complete 11-point sweep in {unit}. The paper-facing figures use only CF1=1, 4, 16, 64, and 256.",
        16,
    )
    by = {(row["case"], row["operation"], row["backend"]): row for row in rows}
    output = []
    for case in CLUSTER_ORDER:
        for operation in OPERATIONS:
            cr = by[(case, operation, "CRoaring")]
            values = []
            for backend in BACKENDS:
                row = by[(case, operation, backend)]
                values.append(float(row["throughput_ops_s"] if throughput else row["time_ms"]))
            output.append([
                case, float(cr["cf1"]), "Yes" if case in PLOTTED_CLUSTER else "No", operation,
                *values,
                *[float(by[(case, operation, backend)]["latency_over_ddc"]) for backend in ("CRoaring", "WAH", "EWAH")],
                cr["winner"], int(cr["replicates"]), float(cr["min_ms"]), float(cr["max_ms"]), cr["source_operation"],
            ])
    value_suffix = "op/s" if throughput else "ms"
    headers = [
        "Case", "CF1 actual", "In paper figure?", "Operation",
        f"DDC {value_suffix}", f"CR native {value_suffix}", f"WAH {value_suffix}", f"EWAH {value_suffix}",
        "CR/DDC latency", "WAH/DDC latency", "EWAH/DDC latency", "Winner",
        "CR reps", "CR min ms", "CR max ms", "CR source operation",
    ]
    end, h = write_table(
        ws, 4, headers, output,
        {1: 13, 2: 17, 3: 17, 4: 13, 5: 16, 6: 17, 7: 16, 8: 16,
         9: 17, 10: 17, 11: 18, 12: 14, 13: 12, 14: 15, 15: 15, 16: 20},
        "E5",
    )
    apply_number_format(ws, range(5, end + 1), [2], "#,##0.00")
    value_format = "#,##0.00" if throughput else "0.000000"
    apply_number_format(ws, range(5, end + 1), [5, 6, 7, 8], value_format)
    apply_number_format(ws, range(5, end + 1), [9, 10, 11], "0.000x")
    apply_number_format(ws, range(5, end + 1), [14, 15], "0.000000")
    for row in range(5, end + 1):
        for column in (9, 10, 11):
            ratio_fill(ws.cell(row, column), float(ws.cell(row, column).value))
        if ws.cell(row, h["In paper figure?"]).value == "Yes":
            ws.cell(row, h["In paper figure?"]).fill = PatternFill("solid", fgColor=BLUE)


def matrix_block(ws, start_row: int, title_text: str, labels: tuple[str, ...], values: dict[tuple[str, str], float], ratio: bool) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=11)
    cell = ws.cell(start_row, 1, title_text)
    cell.font = Font(name="Arial", size=14, bold=True, color=TEXT)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    for index, label in enumerate(labels, start=2):
        header = ws.cell(start_row + 1, index, label)
        header.font = HEADER_FONT
        header.fill = PatternFill("solid", fgColor=NAVY)
        header.alignment = Alignment(horizontal="center")
    ws.cell(start_row + 1, 1, "A \\ B").font = HEADER_FONT
    ws.cell(start_row + 1, 1).fill = PatternFill("solid", fgColor=NAVY)
    for row_index, label_a in enumerate(labels, start=start_row + 2):
        ws.cell(row_index, 1, label_a)
        ws.cell(row_index, 1).font = Font(name="Arial", size=12, bold=True)
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
        for column_index, label_b in enumerate(labels, start=2):
            value = values[(label_a, label_b)]
            cell = ws.cell(row_index, column_index, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "0.00x" if ratio else "0.000"
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if ratio:
                ratio_fill(cell, value)
    if not ratio:
        data_range = f"B{start_row + 2}:K{start_row + 11}"
        vals = list(values.values())
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="num", start_value=min(vals), start_color="C6E0B4",
                mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                end_type="num", end_value=max(vals), end_color="F4CCCC",
            ),
        )
    return start_row + 13


def build_density_sheets(wb: Workbook, rows: list[dict[str, str]]) -> None:
    data = {(row["backend"], row["density_A"], row["density_B"]): float(row["or_time_ms"]) for row in rows}
    ws = wb.create_sheet("06 Density Latency")
    title(ws, "Density OR Latency Matrices", "Each backend uses its own color scale. Values are milliseconds; colors are not comparable across matrices.", 11)
    row = 4
    for backend in DENSITY_BACKENDS:
        values = {(a, b): data[(backend, a, b)] for a in DENSITIES for b in DENSITIES}
        row = matrix_block(ws, row, f"{backend} OR latency (ms)", DENSITIES, values, False)
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 12
    for column in range(2, 12):
        ws.column_dimensions[get_column_letter(column)].width = 11

    ratio_ws = wb.create_sheet("07 Density vs DDC")
    title(ratio_ws, "Density OR Latency Relative to DDC", "Green is faster than DDC; pale blue is approximately equal; yellow/red is slower. Fixed semantic bands avoid scale compression by large outliers.", 11)
    row = 4
    for backend in DENSITY_BACKENDS:
        if backend == "DDC":
            continue
        values = {(a, b): data[(backend, a, b)] / data[("DDC", a, b)] for a in DENSITIES for b in DENSITIES}
        row = matrix_block(ratio_ws, row, f"{backend} / DDC", DENSITIES, values, True)
    ratio_ws.freeze_panes = "B5"
    ratio_ws.column_dimensions["A"].width = 12
    for column in range(2, 12):
        ratio_ws.column_dimensions[get_column_letter(column)].width = 11


def build_delivery(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("08 Delivery Impact")
    title(ws, "CRoaring Dense-Delivery Impact", "Native and delivered CR rows were recorded by the same fresh binary. Historical delivered values are retained as drift anchors.", 15)
    headers = [
        "Group", "Case", "Operation", "CR native ms", "CR delivered ms", "Historical delivered ms",
        "Native / delivered", "Speedup removing delivery", "Delivery overhead ms", "Delivery share",
        "Fresh delivered / historical", "Replicates", "Native source", "Delivered source",
    ]
    output = []
    for row in rows:
        output.append([
            row["group"], row["case"], row["operation"], float(row["native_cr_ms"]),
            float(row["same_run_delivered_cr_ms"]), float(row["historical_delivered_cr_ms"]),
            float(row["native_over_same_run_delivered"]), float(row["native_speedup_from_removing_delivery"]),
            float(row["delivery_overhead_ms"]), float(row["delivery_share_pct"]) / 100.0,
            float(row["same_run_delivered_over_historical"]), int(row["replicates"]),
            row["native_source_operation"], row["delivered_source_operation"],
        ])
    end, h = write_table(ws, 4, headers, output, {1: 12, 2: 24, 3: 13, 4: 16, 5: 17, 6: 20, 7: 18, 8: 21, 9: 20, 10: 16, 11: 24, 12: 12, 13: 20, 14: 20}, "D5")
    apply_number_format(ws, range(5, end + 1), [4, 5, 6, 9], "0.000000")
    apply_number_format(ws, range(5, end + 1), [7, 8, 11], "0.000x")
    apply_number_format(ws, range(5, end + 1), [10], "0.0%")
    for row in range(5, end + 1):
        share = float(ws.cell(row, h["Delivery share"]).value)
        ws.cell(row, h["Delivery share"]).fill = PatternFill("solid", fgColor=GREEN if share < 0.25 else YELLOW if share < 0.5 else ORANGE if share < 0.8 else RED)
        drift = abs(float(ws.cell(row, h["Fresh delivered / historical"]).value) - 1.0)
        if drift > 0.10:
            ws.cell(row, h["Fresh delivered / historical"]).fill = PatternFill("solid", fgColor=RED)
        elif drift > 0.05:
            ws.cell(row, h["Fresh delivered / historical"]).fill = PatternFill("solid", fgColor=YELLOW)


def build_raw_sheet(wb: Workbook, name: str, rows: list[dict[str, str]], subtitle: str) -> None:
    ws = wb.create_sheet(name)
    headers = list(rows[0])
    title(ws, name, subtitle, len(headers))
    values = []
    for row in rows:
        converted = []
        for column in headers:
            value = row[column]
            if value == "":
                converted.append(None)
            elif column in RAW_INTEGER_FIELDS:
                converted.append(int(value))
            elif column in RAW_FLOAT_FIELDS:
                converted.append(float(value))
            else:
                converted.append(value)
        values.append(converted)
    end, h = write_table(ws, 4, headers, values, None, "A5")
    for column_name in ("time_ms", "min_ms", "max_ms", "or_time_ms"):
        if column_name in h:
            apply_number_format(ws, range(5, end + 1), [h[column_name]], "0.000000")
    if "latency_over_ddc" in h:
        apply_number_format(ws, range(5, end + 1), [h["latency_over_ddc"]], "0.000x")
    if "source_file" in h:
        ws.column_dimensions[get_column_letter(h["source_file"])].width = 125
    if "case" in h:
        ws.column_dimensions[get_column_letter(h["case"])].width = 27
    if "source_operation" in h:
        ws.column_dimensions[get_column_letter(h["source_operation"])].width = 20
    if "native_source_operation" in h:
        ws.column_dimensions[get_column_letter(h["native_source_operation"])].width = 22
    if "delivered_source_operation" in h:
        ws.column_dimensions[get_column_letter(h["delivered_source_operation"])].width = 22
    if name == "11 Density Raw":
        ws.column_dimensions["A"].width = 16
    for row in range(5, end + 1):
        ws.row_dimensions[row].height = 22
        for column in range(1, len(headers) + 1):
            ws.cell(row, column).alignment = Alignment(vertical="center", wrap_text=False)


def build_figures(wb: Workbook) -> None:
    ws = wb.create_sheet("13 Figures")
    title(ws, "Key Native-CR Figures", "Large embedded previews; full PDF and PNG files remain in the experiment figures directory.", 20)
    specs = (
        ("earth_or_comp_native_relative.png", "A4", 760),
        ("job_comp_native_relative.png", "A23", 760),
        ("comp_clustering_native.png", "A42", 560),
        ("density_grid_croaring_native.png", "K42", 560),
    )
    for filename, anchor, width in specs:
        path = FIGURES / filename
        image = Image(path)
        scale = width / image.width
        image.width = width
        image.height = image.height * scale
        ws.add_image(image, anchor)
    for column in range(1, 22):
        ws.column_dimensions[get_column_letter(column)].width = 10
    ws.sheet_view.zoomScale = 80


def build_sources(wb: Workbook, sources: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("14 Sources")
    title(ws, "Source and Provenance Hashes", "SHA-256 values cover formal raw CSVs, logs, metadata, frozen baselines, and analysis inputs. Protocol-mismatch runs are excluded.", 3)
    rows = [[index, item["path"], item["sha256"]] for index, item in enumerate(sources, start=1)]
    end, _ = write_table(ws, 4, ["#", "Path", "SHA-256"], rows, {1: 8, 2: 125, 3: 82}, "B5")
    for row in range(5, end + 1):
        ws.row_dimensions[row].height = 22
        for column in range(1, 4):
            ws.cell(row, column).alignment = Alignment(vertical="center", wrap_text=False)


def load_data() -> dict[str, object]:
    earth = read_csv(RESULTS / "earth_native_logical.csv")
    job = read_csv(RESULTS / "job_native_logical.csv")
    cluster = read_csv(RESULTS / "cluster_native_logical.csv")
    density = read_csv(RESULTS / "density_native_heatmap.csv")
    anchors = read_csv(RESULTS / "anchor_comparison.csv")
    table4_run = {row["slug"]: row for row in read_csv(RESULTS / "table4_cr_run_sizes.csv") if row["slug"] != "TOTAL"}
    summary = json.loads((RESULTS / "summary.json").read_text(encoding="utf-8"))

    feature_path = WORKSPACE / "R3W1_realworld_excel_20260809/job_unified10/job-unified10-features.csv"
    size_path = WORKSPACE / "R3W1_realworld_excel_20260809/job_unified10/job-unified10-size.csv"
    features = {row["field"]: row for row in read_csv(feature_path)}
    sizes = read_csv(size_path)
    table4_rows = []
    total = defaultdict(float)
    for size in sizes:
        slug = size["field"]
        feature = features[slug]
        run = table4_run[slug.replace(".", "_")]
        backend_bytes = {
            "DDC": int(size["ddc_payload_bits"]) / 8.0,
            "WAH": int(size["wah_payload_bytes"]),
            "EWAH": int(size["ewah_payload_bytes"]),
            "CRoaring": int(run["cr_run_payload_bytes"]),
        }
        for backend, value in backend_bytes.items():
            total[backend] += value
        winner = min(backend_bytes, key=backend_bytes.get)
        cr_plain = int(run["cr_plain_payload_bytes"])
        cr_run = int(run["cr_run_payload_bytes"])
        ddc = backend_bytes["DDC"]
        table4_rows.append([
            size["source"], int(size["n"]), int(size["c"]), float(feature["avg_density"]),
            float(feature["avg_1_run"]), float(feature["gini"]), ddc / (1 << 20),
            backend_bytes["WAH"] / (1 << 20), backend_bytes["EWAH"] / (1 << 20),
            cr_plain / (1 << 20), cr_run / (1 << 20), cr_run,
            1.0 - cr_run / cr_plain, cr_run / ddc, cr_run / ddc - 1.0,
            winner, "Fresh CR plain/run; frozen DDC/WAH/EWAH",
        ])
    table4_rows.append([
        "TOTAL", "", "", "", "", "", total["DDC"] / (1 << 20), total["WAH"] / (1 << 20),
        total["EWAH"] / (1 << 20), sum(int(row["cr_plain_payload_bytes"]) for row in table4_run.values()) / (1 << 20),
        total["CRoaring"] / (1 << 20), int(total["CRoaring"]),
        1.0 - total["CRoaring"] / sum(int(row["cr_plain_payload_bytes"]) for row in table4_run.values()),
        total["CRoaring"] / total["DDC"], total["CRoaring"] / total["DDC"] - 1.0,
        min(total, key=total.get), "Aggregate",
    ])

    anchor_map = {(row["group"], row["case"], row["operation"]): row for row in anchors}
    earth_map = {(row["case"], row["operation"], row["backend"]): row for row in earth}
    job_map = {(row["case"], row["operation"], row["backend"]): row for row in job}
    earth_subset = [float(earth_map[(case, op, "CRoaring")]["latency_over_ddc"]) for case in EARTH_ORDER for op in ("OR", "COMP")]
    earth_old = [
        float(anchor_map[("earth", case, op)]["historical_delivered_cr_ms"]) /
        float(earth_map[(case, op, "DDC")]["time_ms"])
        for case in EARTH_ORDER for op in ("OR", "COMP")
    ]
    job_subset = [float(job_map[(case, "COMP", "CRoaring")]["latency_over_ddc"]) for case in JOB_ORDER]
    job_old = [
        float(anchor_map[("job", case, "COMP")]["historical_delivered_cr_ms"]) /
        float(job_map[(case, "COMP", "DDC")]["time_ms"])
        for case in JOB_ORDER
    ]
    density_map = {(row["backend"], row["density_A"], row["density_B"]): float(row["or_time_ms"]) for row in density}
    density_new = [density_map[("CRoaring", a, b)] / density_map[("DDC", a, b)] for a in DENSITIES for b in DENSITIES]
    density_counts = (66, 328, 655, 1311, 3277, 6554, 13107, 19661, 26214, 32768)
    density_old = []
    for i, count_a in enumerate(density_counts):
        for j, count_b in enumerate(density_counts):
            low, high = sorted((count_a, count_b))
            historical = float(
                anchor_map[("density", f"A{low}_B{high}", "OR")][
                    "historical_delivered_cr_ms"
                ]
            )
            density_old.append(historical / density_map[("DDC", DENSITIES[i], DENSITIES[j])])
    cluster_ratios = [float(row["latency_over_ddc"]) for row in cluster if row["backend"] == "CRoaring"]
    cluster_map = {
        (row["case"], row["operation"], row["backend"]): row for row in cluster
    }
    cluster_old = [
        float(anchor_map[("cluster", case, operation)]["historical_delivered_cr_ms"])
        / float(cluster_map[(case, operation, "DDC")]["time_ms"])
        for case in CLUSTER_ORDER
        for operation in OPERATIONS
    ]
    overview_metrics = [
        ("Table 4 CR run total", f"{total['CRoaring'] / (1 << 20):.3f} MiB", f"{sum(int(row['cr_plain_payload_bytes']) for row in table4_run.values()) / (1 << 20):.3f} MiB plain", "Run optimization reduces CR payload by 24.93%; CR run remains 25.03% above DDC."),
        ("Earth OR+COMP CR/DDC geometric mean", f"{geometric_mean(earth_subset):.4f}x", f"{geometric_mean(earth_old):.4f}x", "CR wins 5 of 10 figure-facing cells under the native boundary."),
        ("JOB COMP CR/DDC geometric mean", f"{geometric_mean(job_subset):.4f}x", f"{geometric_mean(job_old):.4f}x", "Large reduction, but frozen DDC still uses a shorter equivalent COMP plan."),
        ("Cluster all-operation CR/DDC geometric mean", f"{geometric_mean(cluster_ratios):.4f}x", f"{geometric_mean(cluster_old):.4f}x", "Complete 11-point sweep; 5 points are shown in the current figure."),
        ("Density OR CR/DDC geometric mean", f"{geometric_mean(density_new):.4f}x", f"{geometric_mean(density_old):.4f}x", "100-cell symmetric matrix derived from 55 independent upper-triangle measurements."),
    ]
    return {
        "earth": earth,
        "job": job,
        "cluster": cluster,
        "density": density,
        "anchors": anchors,
        "anchor_map": anchor_map,
        "table4_rows": table4_rows,
        "table4_totals": {
            "ddc": total["DDC"] / (1 << 20),
            "plain": sum(int(row["cr_plain_payload_bytes"]) for row in table4_run.values()) / (1 << 20),
            "run": total["CRoaring"] / (1 << 20),
        },
        "overview_metrics": overview_metrics,
        "sources": summary["sources"],
    }


data_global: dict[str, object] = {}


def build_workbook() -> None:
    global data_global
    data_global = load_data()
    workbook = Workbook()
    workbook.remove(workbook.active)
    build_overview(workbook, data_global)
    build_table4(workbook, data_global)
    logical_analysis_sheet(workbook, "02 Earth Analysis", data_global["earth"], EARTH_ORDER)
    logical_analysis_sheet(workbook, "03 JOB Analysis", data_global["job"], JOB_ORDER)
    build_cluster(workbook, data_global["cluster"], False)
    build_cluster(workbook, data_global["cluster"], True)
    build_density_sheets(workbook, data_global["density"])
    build_delivery(workbook, data_global["anchors"])
    build_raw_sheet(workbook, "09 Earth JOB Raw", data_global["earth"] + data_global["job"], "Compact source rows. CR is fresh; non-CR rows are frozen baselines.")
    build_raw_sheet(workbook, "10 Cluster Raw", data_global["cluster"], "Complete 11-point compact source table with authoritative actual CF1 values.")
    build_raw_sheet(workbook, "11 Density Raw", data_global["density"], "600 plotted cells; CR lower triangle is mirrored from the 55 measured upper-triangle combinations.")
    build_raw_sheet(workbook, "12 Anchor Raw", data_global["anchors"], "Fresh native/delivered CR measurements and historical delivered anchors.")
    build_figures(workbook)
    build_sources(workbook, data_global["sources"])
    workbook.properties.creator = "Native CR sensitivity experiment"
    workbook.properties.title = "Native CRoaring no-BTV analysis"
    workbook.properties.description = "Readable analysis workbook generated from the isolated experiment branch."
    workbook.save(OUTPUT)


def validate_workbook() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    expected = [
        "00 Overview", "01 Table4 Run Size", "02 Earth Analysis", "03 JOB Analysis",
        "04 Cluster Latency", "05 Cluster Throughput", "06 Density Latency",
        "07 Density vs DDC", "08 Delivery Impact", "09 Earth JOB Raw",
        "10 Cluster Raw", "11 Density Raw", "12 Anchor Raw", "13 Figures", "14 Sources",
    ]
    if workbook.sheetnames != expected:
        raise RuntimeError(f"unexpected sheets: {workbook.sheetnames}")
    checks = {
        "01 Table4 Run Size": (15, 17),
        "02 Earth Analysis": (24, 18),
        "03 JOB Analysis": (44, 18),
        "04 Cluster Latency": (48, 16),
        "05 Cluster Throughput": (48, 16),
        "09 Earth JOB Raw": (244, 13),
        "10 Cluster Raw": (180, 15),
        "11 Density Raw": (604, 4),
        "12 Anchor Raw": (163, 14),
        "14 Sources": (302, 3),
    }
    for sheet, (rows, columns) in checks.items():
        ws = workbook[sheet]
        if ws.max_row != rows or ws.max_column != columns:
            raise RuntimeError(f"{sheet}: {ws.max_row}x{ws.max_column} != {rows}x{columns}")
    table = workbook["01 Table4 Run Size"]
    if abs(table.cell(15, 11).value - 218.62847709655762) > 1e-12:
        raise RuntimeError("Table 4 CR run total mismatch")
    if workbook["13 Figures"]._images.__len__() != 4:
        raise RuntimeError("expected four embedded figures")
    for sheet in ("09 Earth JOB Raw", "10 Cluster Raw", "11 Density Raw", "12 Anchor Raw"):
        ws = workbook[sheet]
        header = {ws.cell(4, column).value: column for column in range(1, ws.max_column + 1)}
        for field in (RAW_INTEGER_FIELDS | RAW_FLOAT_FIELDS) & set(header):
            value = ws.cell(5, header[field]).value
            if not isinstance(value, (int, float)):
                raise RuntimeError(f"{sheet}: {field} is not numeric")
    for ws in workbook.worksheets:
        if ws.sheet_view.zoomScale is None or ws.sheet_view.zoomScale < 80:
            raise RuntimeError(f"{ws.title}: unreadable zoom")


def main() -> None:
    build_workbook()
    validate_workbook()
    print(f"wrote {OUTPUT}")


if __name__ == "__main__":
    main()
