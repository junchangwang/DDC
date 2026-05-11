#!/usr/bin/env python3
"""Generate per-density Excel report matching the user's previous format.

Layout per density block:
  row 1 : "Rows: N"
  row 2 : "Density: X | Loaded: c"  + storage info per backend column
  row 3 : (blank)
  row 4 : YELLOW "只测operation-only的pre-alloc"  + backend headers
  row 5 : Loaded                    (load times)
  row 6 : bitOr                     (op-only OR_op times)
  row 7 : bitAnd                    (op-only AND_op times)
  row 8 : bitXor                    (op-only XOR_op times)
  row 9 : Multi-way OR              (multi-OR times — same value, no op-only mode)
  row 10: (blank)
  row 11: "Rows: N"
  row 12: "Density: X | Loaded: c"
  row 13: BLUE "正常操作[Summary]"      + backend headers
  row 14: Loaded
  row 15: bitOr                     (pairwise OR times, with handle alloc)
  row 16: bitAnd
  row 17: bitXor
  row 18: Multi-way OR
  row 19: (blank between density blocks)
"""
from __future__ import annotations

import csv
import re
import statistics
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip3 install openpyxl")
    raise SystemExit(1)

ROOT = Path(__file__).parent
CSV_PATH = ROOT / "results_full.csv"
STORAGE_DIR = ROOT / "storage_info"
OUT_XLSX = ROOT / "results_report.xlsx"
ROWS = 100_000_000

# Backends in display order — matches user's previous Excel screenshot.
# basic / AVX512 / Croaring-bitvector left blank (no data for those here).
# Only Croaring and Combit-L4 populate the storage row; others show only
# timing data.
BACKEND_DISPLAY = ["basic", "AVX512", "WAH", "EWAH", "Concise",
                   "Croaring-bitvector", "Croaring", "Combit-L4"]

# Map CSV "backend" name → display column name.
CSV_TO_DISPLAY = {
    "WAH (FastBit)":  "WAH",
    "EWAH":           "EWAH",
    "Concise":        "Concise",
    "CRoaring":       "Croaring",
    "ComBIT (New)":   "Combit-L4",
}
# Map display name → which subdir prefix to look up storage logs.
DISPLAY_TO_ALGO = {
    "WAH":         "wah",
    "EWAH":        "ewah",
    "Concise":     "concise",
    "Croaring":    "roaring",
    "Combit-L4":   "combit",
}

# Auto-detect cards from CSV.
def discover_cards():
    cards = set()
    with CSV_PATH.open() as f:
        for row in csv.DictReader(f):
            cards.add(int(row["cardinality"]))
    return sorted(cards)


# ----------------------------------------------------------------
# Load timings
# ----------------------------------------------------------------

def load_timings():
    times = defaultdict(list)
    cards_seen = set()
    with CSV_PATH.open() as f:
        for row in csv.DictReader(f):
            key = (row["backend"], int(row["cardinality"]), row["operation"])
            times[key].append(float(row["time_ms"]))
            cards_seen.add(int(row["cardinality"]))
    # multi-OR is multi-second per iter and very sensitive to CPU contention
    # from concurrent processes.  When a noisy iteration spikes the time
    # 2-3x, median picks the polluted value.  Use min() for multi-OR (the
    # uncontaminated baseline) while keeping median for sub-millisecond ops
    # where short-burst noise averages out cleanly.
    aggregates = {}
    for k, v in times.items():
        if k[2] == "multi-OR":
            aggregates[k] = min(v)
        else:
            aggregates[k] = statistics.median(v)
    return aggregates, sorted(cards_seen)


# ----------------------------------------------------------------
# Parse storage logs
# ----------------------------------------------------------------

# Parse the [Breakdown] line — the [ComBit Storage] line is buggy
# (l3_bytes is wildly inflated at high cards, e.g. 9313 MB when total
# is 1450 MB).  The Breakdown line is computed from a separate code
# path and matches actual on-disk sizes.
COMBIT_RE = re.compile(
    r"\[Breakdown\]\s+ComBit\s+"
    r"L1=([\d.]+)\s*MiB\s+"
    r"L2=([\d.]+)\s*MiB\s+"
    r"L3=([\d.]+)\s*MiB\s+"
    r"L4=([\d.]+)\s*MiB\s+"
    r"total=([\d.]+)\s*MiB"
)
CR_RE = re.compile(
    r"\[CRoaring Storage\]\s+array=(\d+)\s*\(([\d.]+)\s*MB\)"
    r"\s*\|\s*run=(\d+)\s*\(([\d.]+)\s*MB\)"
    r"\s*\|\s*bitset=(\d+)\s*\(([\d.]+)\s*MB\)"
    r"\s*\|\s*total_bytes=(\d+)"
)
COMPRESSED_RE = re.compile(r"Total compressed:\s+(\d+)\s+bytes")
# Best-effort bytes for ewah/concise/wah — the Storage line just shows
# total_compressed in the [Storage] block.
TOTAL_COMP_RE = re.compile(r"compressed_bytes:\s+(\d+)\s+bytes")


def parse_storage_log(log: Path):
    out = {}
    if not log.exists():
        return out
    text = log.read_text()
    m = COMPRESSED_RE.search(text)
    if m:
        out["compressed_bytes"] = int(m.group(1))
    m = COMBIT_RE.search(text)
    if m:
        out["combit_l1_MB"]    = float(m.group(1))
        out["combit_l2_MB"]    = float(m.group(2))
        out["combit_l3_MB"]    = float(m.group(3))
        out["combit_l4_MB"]    = float(m.group(4))
        out["combit_total_MB"] = float(m.group(5))
    m = CR_RE.search(text)
    if m:
        out["cr_array_n"]  = int(m.group(1))
        out["cr_array_MB"] = float(m.group(2))
        out["cr_run_n"]    = int(m.group(3))
        out["cr_run_MB"]   = float(m.group(4))
        out["cr_bitset_n"] = int(m.group(5))
        out["cr_bitset_MB"] = float(m.group(6))
        out["cr_total_bytes"] = int(m.group(7))
    return out


def collect_storage(cards):
    storage = {}
    for c in cards:
        for disp, algo in DISPLAY_TO_ALGO.items():
            log = STORAGE_DIR / f"c{c}_{algo}.log"
            storage[(c, disp)] = parse_storage_log(log)
    return storage


# ----------------------------------------------------------------
# Storage-cell content per backend
# ----------------------------------------------------------------

def storage_cell(disp: str, st: dict) -> str:
    if not st:
        return ""
    cb = st.get("compressed_bytes")
    if disp == "Combit-L4":
        return (f"L4: {st.get('combit_l4_MB',0):.4f} MB  |  "
                f"L3: {st.get('combit_l3_MB',0):.4f} MB  |  "
                f"L2: {st.get('combit_l2_MB',0):.4f} MB  |  "
                f"L1: {st.get('combit_l1_MB',0):.4f} MB  |  "
                f"total in-memory: {st.get('combit_total_MB',0):.4f} MB"
                + (f"  |  on-disk: {cb} bytes ({cb/(1024*1024):.2f} MB)" if cb else ""))
    if disp == "Croaring":
        # Each n_* is the COUNT OF CONTAINERS of that type, not elements.
        # Each MB value is sum of element bytes for those containers.
        # For arrays: cardinality * 2 bytes per container, summed.
        # For bitsets: 8192 bytes per container.  For runs: 4 bytes per run.
        # total_bytes = bytes_arr + bytes_bs + bytes_run (in-memory size).
        # compressed_bytes (separate) = serialized file size on disk
        # (includes per-container headers).
        arr_n = st.get("cr_array_n", 0)
        run_n = st.get("cr_run_n", 0)
        bs_n  = st.get("cr_bitset_n", 0)
        arr_MB = st.get("cr_array_MB", 0)
        run_MB = st.get("cr_run_MB", 0)
        bs_MB  = st.get("cr_bitset_MB", 0)
        cr_total = st.get("cr_total_bytes", 0)
        total_MB = cr_total / (1024 * 1024) if cr_total else (arr_MB + run_MB + bs_MB)
        return (f"array: {arr_n} containers ({arr_MB:.2f} MB)  |  "
                f"run: {run_n} ({run_MB:.2f} MB)  |  "
                f"bitset: {bs_n} containers ({bs_MB:.2f} MB)  |  "
                f"total in-memory: {total_MB:.2f} MB"
                + (f"  |  on-disk: {cb} bytes ({cb/(1024*1024):.2f} MB)" if cb else ""))
    return ""


# ----------------------------------------------------------------
# Excel building
# ----------------------------------------------------------------

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
BLUE   = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
HEADER_FONT = Font(bold=True)


def fmt_ms(t):
    return f"{t:.4f} ms" if t is not None else ""


def get_card_for_op(medians, csv_b, c, op):
    """Find result_cardinality for the given (backend, card, op)."""
    # Need a separate pass — we don't track it in `medians`.
    return None


def build_excel(medians, cards, storage):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Per-density blocks
    row = 1
    for c in cards:
        density = 1.0 / c

        # ── Op-only section ─────────────────────────────
        ws.cell(row=row, column=1, value=f"Rows: {ROWS}")
        row += 1

        ws.cell(row=row, column=1,
                value=f"Density: {density:.5g} | Loaded: {c}")
        for col_idx, disp in enumerate(BACKEND_DISPLAY, start=2):
            content = storage_cell(disp, storage.get((c, disp), {}))
            cell = ws.cell(row=row, column=col_idx, value=content)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60
        row += 1

        # Yellow header
        cell = ws.cell(row=row, column=1, value="只测operation-only的pre-alloc")
        cell.fill = YELLOW
        cell.font = HEADER_FONT
        for col_idx, disp in enumerate(BACKEND_DISPLAY, start=2):
            c2 = ws.cell(row=row, column=col_idx, value=disp)
            c2.fill = YELLOW
            c2.font = HEADER_FONT
        row += 1

        op_only_rows = [
            ("Loaded",       "load"),
            ("bitOr",        "OR_op"),
            ("bitAnd",       "AND_op"),
            ("bitXor",       "XOR_op"),
            ("Multi-way OR", "multi-OR"),
        ]
        for label, op in op_only_rows:
            ws.cell(row=row, column=1, value=label)
            for col_idx, disp in enumerate(BACKEND_DISPLAY, start=2):
                csv_b = next((k for k, v in CSV_TO_DISPLAY.items() if v == disp), None)
                if csv_b is None:
                    continue
                t = medians.get((csv_b, c, op))
                if t is not None:
                    ws.cell(row=row, column=col_idx, value=fmt_ms(t))
            row += 1

        # ── Summary section ─────────────────────────────
        ws.cell(row=row, column=1, value=f"Rows: {ROWS}")
        row += 1

        ws.cell(row=row, column=1,
                value=f"Density: {density:.5g} | Loaded: {c}")
        row += 1

        cell = ws.cell(row=row, column=1, value="正常操作[Summary]")
        cell.fill = BLUE
        cell.font = HEADER_FONT
        for col_idx, disp in enumerate(BACKEND_DISPLAY, start=2):
            c2 = ws.cell(row=row, column=col_idx, value=disp)
            c2.fill = BLUE
            c2.font = HEADER_FONT
        row += 1

        summary_rows = [
            ("Loaded",       "load"),
            ("bitOr",        "OR"),
            ("bitAnd",       "AND"),
            ("bitXor",       "XOR"),
            ("Multi-way OR", "multi-OR"),
        ]
        for label, op in summary_rows:
            ws.cell(row=row, column=1, value=label)
            for col_idx, disp in enumerate(BACKEND_DISPLAY, start=2):
                csv_b = next((k for k, v in CSV_TO_DISPLAY.items() if v == disp), None)
                if csv_b is None:
                    continue
                t = medians.get((csv_b, c, op))
                if t is not None:
                    ws.cell(row=row, column=col_idx, value=fmt_ms(t))
            row += 1
        # No blank row between density blocks — keep tightly packed
        # so the user can paste cleanly without empty-row gaps.

    # Column widths
    ws.column_dimensions["A"].width = 36
    for col_idx in range(2, len(BACKEND_DISPLAY) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 32

    wb.save(OUT_XLSX)
    print(f"  Wrote {OUT_XLSX}")


# ----------------------------------------------------------------
# Main
# ----------------------------------------------------------------

if __name__ == "__main__":
    print("=== Loading timings ===")
    medians, cards = load_timings()
    print(f"  cards: {cards}")

    print("=== Parsing storage logs ===")
    storage = collect_storage(cards)
    have_storage = sum(1 for v in storage.values() if v)
    print(f"  parsed {have_storage}/{len(storage)} non-empty records")

    print("=== Generating Excel ===")
    build_excel(medians, cards, storage)
