#!/usr/bin/env python3
"""update_results_report.py

Reads `build/refresh_results.csv` (the bench output for WAH/EWAH/Concise/
Bitset/Bitset_AVX512) and overwrites the corresponding cells in
`results_report.xlsx` — leaving Croaring, Croaring-bitvector, DDC-L4
columns untouched.

Layout (per density block):
  row N+0:  "Rows: 100000000"
  row N+1:  "Density: X.X | Loaded: N"
  row N+2:  "只测operation-only的pre-alloc"   ← op-only header
  row N+3:  Loaded                          ← load_ms
  row N+4:  bitOr                           ← OR_op_ms (no build)
  row N+5:  bitAnd                          ← AND_op_ms
  row N+6:  bitXor                          ← XOR_op_ms
  row N+7:  Multi-way OR                    ← multi-OR_ms
  row N+8:  "Rows: 100000000"
  row N+9:  "Density: X.X | Loaded: N"
  row N+10: "正常操作[Summary]"               ← pair (with build) header
  row N+11: Loaded                          ← load_ms (same)
  row N+12: bitOr                           ← OR_ms (with build)
  row N+13: bitAnd                          ← AND_ms
  row N+14: bitXor                          ← XOR_ms
  row N+15: Multi-way OR                    ← multi-OR_ms

Columns:
  B(2) = basic              ← Bitset (Plain)
  C(3) = AVX512             ← Bitset (AVX512)
  D(4) = WAH                ← WAH (FastBit)
  E(5) = EWAH               ← EWAH
  F(6) = Concise            ← Concise
  G(7) = Croaring-bitvector ← (skip)
  H(8) = Croaring           ← (skip, already filled)
  I(9) = DDC-L4          ← (skip, already filled)
"""
from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path
from statistics import median

import openpyxl
from openpyxl.styles import Alignment


# Map CSV "backend" name → Excel column index (1-based).
BACKEND_TO_COL = {
    "Bitset (Plain)":  2,   # B
    "Bitset (AVX512)": 3,   # C
    "WAH (FastBit)":   4,   # D
    "EWAH":            5,   # E
    "Concise":         6,   # F
}

# Operation labels appearing in column A of the data rows.
OP_LABEL_TO_KEY_OPONLY = {
    "Loaded":       "load",
    "bitOr":        "OR_op",
    "bitAnd":       "AND_op",
    "bitXor":       "XOR_op",
    "Multi-way OR": "multi-OR",
}
OP_LABEL_TO_KEY_SUMMARY = {
    "Loaded":       "load",
    "bitOr":        "OR",
    "bitAnd":       "AND",
    "bitXor":       "XOR",
    "Multi-way OR": "multi-OR",
}


# ============================================================
# 1) Parse refresh_results.csv → (backend, card, op) → median ms,
#    plus (backend, card) → total compressed_bytes
# ============================================================
def parse_results(path: Path):
    times: dict[tuple[str, int, str], list[float]] = defaultdict(list)
    bytes_at: dict[tuple[str, int], int] = {}

    with path.open() as f:
        r = csv.DictReader(f)
        for row in r:
            be   = row["backend"]
            card = int(row["cardinality"])
            op   = row["operation"]
            t_ms = float(row["time_ms"])
            cb   = int(row["compressed_bytes"])

            times[(be, card, op)].append(t_ms)
            if cb > 0:
                bytes_at[(be, card)] = cb

    medians = {k: median(v) for k, v in times.items()}
    return medians, bytes_at


# ============================================================
# 2) Walk the spreadsheet, identify each density block, fill cells
# ============================================================
DENSITY_RE = re.compile(r"Density:\s*([0-9.]+)\s*\|\s*Loaded:\s*(\d+)")


def find_density_blocks(ws):
    """Yield (top_row, density, cardinality) for each '只测...' header block."""
    last_card = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        m = DENSITY_RE.search(str(v))
        if m:
            last_card = int(m.group(2))
        if "只测" in str(v) and last_card is not None:
            # The header row IS r; the 5 op rows follow at r+1..r+5.
            yield r, last_card, "oponly"
        if "正常" in str(v) and last_card is not None:
            yield r, last_card, "summary"


def write_cell(ws, r, c, ms):
    cell = ws.cell(r, c)
    cell.value = f"{ms:.4f} ms"
    cell.alignment = Alignment(horizontal="center", vertical="center")


def main():
    csv_in   = Path("build/refresh_results.csv")
    xlsx_io  = Path("results_report.xlsx")

    medians, bytes_at = parse_results(csv_in)
    print(f"[load] {len(medians)} timing medians, {len(bytes_at)} (backend,card) memory entries")

    wb = openpyxl.load_workbook(xlsx_io)
    ws = wb.active

    cells_written = 0
    for hdr_row, card, kind in find_density_blocks(ws):
        op_map = OP_LABEL_TO_KEY_OPONLY if kind == "oponly" else OP_LABEL_TO_KEY_SUMMARY
        # Op rows at hdr_row+1 .. hdr_row+5 — verify by label
        for off in range(1, 6):
            r = hdr_row + off
            label = ws.cell(r, 1).value
            if label not in op_map:
                continue
            key = op_map[label]
            for be, col in BACKEND_TO_COL.items():
                med = medians.get((be, card, key))
                if med is None:
                    continue
                write_cell(ws, r, col, med)
                cells_written += 1

    # Memory rows: find each "Density: ..." row and update memory in cols 2..6.
    # Original format used col I for DDC; we'll add a one-line summary at each
    # memory anchor row (col 2..6) showing "total: N MB" for the matching backend.
    mem_rows_written = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        m = DENSITY_RE.search(str(v))
        if not m:
            continue
        card = int(m.group(2))
        for be, col in BACKEND_TO_COL.items():
            cb = bytes_at.get((be, card))
            if not cb:
                continue
            mb = cb / 1e6
            text = f"total: {cb:,} bytes ({mb:,.2f} MB)"
            cell = ws.cell(r, col)
            cell.value = text
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
            mem_rows_written += 1

    print(f"[fill] {cells_written} timing cells, {mem_rows_written} memory cells written")

    wb.save(xlsx_io)
    print(f"[ok]   saved {xlsx_io}")


if __name__ == "__main__":
    main()
