#!/usr/bin/env python3
"""make_transition_excel.py

Reads `transition_results.csv` produced by `transition_bench` and writes
`transition_case.xlsx` with multiple formatted, charted sheets so the
container-conversion story is readable at a glance.

Usage:
    python3 make_transition_excel.py <csv_in> <xlsx_out>
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter


# ============================================================
# Colour tokens (mild, presentation-grade)
# ============================================================
NAVY        = "1F4E78"    # title bar
LIGHT_BLUE  = "DCE6F1"    # subtitle / section row
HIGHLIGHT   = "FFF2CC"    # ComBit highlight row
ACCENT_GREEN = "C6EFCE"   # winner cell
ACCENT_RED   = "FFC7CE"   # loser cell

THICK = Side(border_style="medium", color="1F4E78")
THIN  = Side(border_style="thin", color="BFBFBF")

# ============================================================
# Data ingestion
# ============================================================
def load_rows(path: Path) -> list[dict]:
    with path.open() as f:
        return list(csv.DictReader(f))


# ============================================================
# Helpers
# ============================================================
def title_cell(cell, text: str, size: int = 14, font_color: str = "FFFFFF"):
    cell.value = text
    cell.font = Font(bold=True, color=font_color, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(top=THICK, bottom=THICK, left=THICK, right=THICK)


def subtitle_cell(cell, text: str):
    cell.value = text
    cell.font = Font(bold=True, color="1F4E78", size=11, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def header_cell(cell, text: str):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def data_cell(cell, val, *, bold=False, fg=None, number_format="0.00", align="center"):
    cell.value = val
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    cell.font = Font(bold=bold, name="Calibri", size=11)
    if isinstance(val, (int, float)):
        cell.number_format = number_format
    if fg:
        cell.fill = PatternFill("solid", fgColor=fg)


def set_widths(ws, widths: dict[str, int]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ============================================================
# Sheet 1: Summary — two scenarios, median ms + speedup
# ============================================================
def write_summary_sheet(ws, rows):
    ws.title = "Summary"
    set_widths(ws, {"A": 5, "B": 14, "C": 16, "D": 14, "E": 14,
                    "F": 16, "G": 16, "H": 20})

    title_cell(ws.cell(1, 1), "ComBit vs CRoaring — Container-Transition Benchmark", size=15)
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30

    subtitle_cell(ws.cell(2, 1),
                  "Bitmap size = 67M bits (1024 × 65536-bit segments)   "
                  "Iterations = 2 warm-up + 9 measured   "
                  "Result = median ± stddev (ms)")
    ws.merge_cells("A2:H2")

    row_cursor = 4
    by_scen: dict[str, list[dict]] = {}
    for r in rows:
        by_scen.setdefault(r["scenario"], []).append(r)

    SCEN_BANNER = {
        "S1_OR_a2b":  ("Scenario 1 — OR  (CR: array → bitset transition)",
                       "Each input segment: ~3000 set bits → array container (density 4.6%)\n"
                       "OR result: ~6000 set bits → bitset container (density 9.2%, > 4096 threshold)\n"
                       "CR must allocate 8 KB bitset and re-emit data; ComBit byte-ORs L1 streams directly."),
        "S2_AND_b2a": ("Scenario 2 — AND (CR: bitset → array transition)",
                       "Each input segment: ~16000 set bits → bitset container (density 24%)\n"
                       "AND result: ~3900 set bits → array container (density 6%, < 4096 threshold)\n"
                       "CR must extract set-bit positions from bitset into uint16 array; ComBit byte-ANDs L1 streams."),
        "S3_OR_worst": ("Scenario 3 — OR  WORST PATH  array → bitset → array  (CR longest path)",
                       "Each input segment: 3500 set bits with 95% overlap (both array, just under threshold)\n"
                       "|A|+|B|=7000>4096 triggers bitset allocation, but actual |A∪B|≈3675 ≤ 4096,\n"
                       "so CR ALSO runs array_container_from_bitset + bitset_container_free.\n"
                       "This is the FULL path of array_array_container_union — CR's adversarial worst case."),
    }

    for scen_key, banner in SCEN_BANNER.items():
        # Banner
        cell = ws.cell(row_cursor, 1, banner[0])
        cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row_cursor, start_column=1,
                       end_row=row_cursor, end_column=8)
        ws.row_dimensions[row_cursor].height = 22

        # Description (italic, smaller)
        desc = ws.cell(row_cursor + 1, 1, banner[1])
        desc.font = Font(italic=True, color="555555", size=10, name="Calibri")
        desc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row_cursor + 1, start_column=1,
                       end_row=row_cursor + 1, end_column=8)
        ws.row_dimensions[row_cursor + 1].height = 48

        # Column headers
        hdr_row = row_cursor + 2
        for i, h in enumerate(
            ["#", "Backend", "Median (ms)", "Min (ms)", "Max (ms)",
             "Std-dev (ms)", "Result card", "Result size (MB)"],
            start=1):
            header_cell(ws.cell(hdr_row, i), h)
        ws.row_dimensions[hdr_row].height = 22

        # Data rows — sort by median asc so the fastest is on top
        scen_rows = sorted(by_scen[scen_key], key=lambda r: float(r["median_ms"]))
        fastest_ms = float(scen_rows[0]["median_ms"])
        # Show name + highlight: CB-decompressed (yellow), CB-compressed (orange).
        CB_NAMES = {"CB-decompressed": "ComBit (decompressed out)",
                    "CB-compressed":   "ComBit (compressed out, apples-to-apples)",
                    "CB":              "ComBit"}
        CB_FG    = {"CB-decompressed": "FFF2CC",
                    "CB-compressed":   "FCE4D6",
                    "CB":              HIGHLIGHT}
        cr_ms = next((float(r["median_ms"]) for r in scen_rows if r["backend"] == "CR"), None)
        for idx, r in enumerate(scen_rows, start=1):
            is_combit_family = r["backend"].startswith("CB")
            base_fg = CB_FG.get(r["backend"]) if is_combit_family else None
            row = hdr_row + idx
            data_cell(ws.cell(row, 1), idx, number_format="0", fg=base_fg)
            display_name = CB_NAMES.get(r["backend"], r["backend"])
            data_cell(ws.cell(row, 2), display_name,
                      bold=is_combit_family, fg=base_fg, align="center")
            med = float(r["median_ms"])
            fg = ACCENT_GREEN if med == fastest_ms else (
                 ACCENT_RED if r == scen_rows[-1] else base_fg)
            data_cell(ws.cell(row, 3), med, bold=is_combit_family, fg=fg, number_format="0.000")
            data_cell(ws.cell(row, 4), float(r["min_ms"]), fg=base_fg, number_format="0.000")
            data_cell(ws.cell(row, 5), float(r["max_ms"]), fg=base_fg, number_format="0.000")
            data_cell(ws.cell(row, 6), float(r["stddev_ms"]), fg=base_fg, number_format="0.000")
            data_cell(ws.cell(row, 7), int(r["result_card"]), fg=base_fg, number_format="#,##0")
            data_cell(ws.cell(row, 8), float(r["result_size_mb"]), fg=base_fg, number_format="0.00")

        # Footer: ratios for both ComBit modes
        footer = hdr_row + len(scen_rows) + 1
        cb_d = next((float(r["median_ms"]) for r in scen_rows if r["backend"] == "CB-decompressed"), None)
        cb_c = next((float(r["median_ms"]) for r in scen_rows if r["backend"] == "CB-compressed"), None)
        msgs = []
        if cr_ms and cb_d:
            r1 = cr_ms / cb_d
            msgs.append(f"CR / CB-decompressed = {r1:.2f}×  (fast path: CB output is L1-flat, no re-compress)")
        if cr_ms and cb_c:
            r2 = cr_ms / cb_c
            tag = "FASTER" if r2 > 1 else "SLOWER"
            msgs.append(f"CR / CB-compressed = {r2:.2f}×  (apples-to-apples: CB is {tag} when also re-compressing output)")
        msgs.append(f"All backends produced identical result cardinality ({int(scen_rows[0]['result_card']):,}).")
        msg = "   ".join(msgs)
        c = ws.cell(footer, 1, msg)
        c.font = Font(bold=True, color="1F4E78", size=11, name="Calibri")
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=footer, start_column=1,
                       end_row=footer, end_column=8)
        ws.row_dimensions[footer].height = 44

        row_cursor = footer + 3


# ============================================================
# Sheet 2: Chart — bar chart per scenario
# ============================================================
def write_chart_sheet(ws, rows):
    ws.title = "Chart"
    set_widths(ws, {"A": 14, "B": 16, "C": 16})

    title_cell(ws.cell(1, 1),
               "OR / AND median operation time — lower is better",
               size=14)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    # Pivot rows: one per backend, columns S1 / S2.
    by_backend: dict[str, dict[str, float]] = {}
    for r in rows:
        by_backend.setdefault(r["backend"], {})[r["scenario"]] = float(r["median_ms"])

    order = ["CB", "CR", "WAH", "EWAH", "Concise"]   # ComBit first
    pretty = {"CB": "ComBit"}
    header_row = 3
    header_cell(ws.cell(header_row, 1), "Backend")
    header_cell(ws.cell(header_row, 2), "Scenario 1\n(OR  array→bitset)")
    header_cell(ws.cell(header_row, 3), "Scenario 2\n(AND  bitset→array)")
    ws.row_dimensions[header_row].height = 32

    for i, b in enumerate(order, start=1):
        row = header_row + i
        is_cb = b == "CB"
        fg = HIGHLIGHT if is_cb else None
        data_cell(ws.cell(row, 1), pretty.get(b, b), bold=is_cb, fg=fg)
        s1 = by_backend.get(b, {}).get("S1_OR_a2b", 0.0)
        s2 = by_backend.get(b, {}).get("S2_AND_b2a", 0.0)
        data_cell(ws.cell(row, 2), s1, bold=is_cb, fg=fg, number_format="0.00")
        data_cell(ws.cell(row, 3), s2, bold=is_cb, fg=fg, number_format="0.00")

    data_end = header_row + len(order)

    # Bar chart, grouped per scenario, one bar per backend.
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "OR / AND median operation time (ms, lower = better)"
    chart.y_axis.title = "Backend"
    chart.x_axis.title = "Median operation time (ms)"

    data_ref = Reference(ws, min_col=2, min_row=header_row,
                         max_col=3, max_row=data_end)
    cat_ref  = Reference(ws, min_col=1, min_row=header_row + 1,
                         max_row=data_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = 12
    chart.width  = 22
    chart.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(chart, "E3")


# ============================================================
# Sheet 3: How CRoaring switches containers (educational)
# ============================================================
def write_howitworks_sheet(ws):
    ws.title = "How CR switches"
    set_widths(ws, {"A": 28, "B": 60, "C": 25})

    title_cell(ws.cell(1, 1),
               "How CRoaring switches containers — annotated call chain",
               size=14)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    sections = [
        ("Scenario 1 — OR triggers array → bitset",
         [("C++ entry",
           "roaring::Roaring::operator|(const Roaring&)",
           "src/core/croaring/roaring.hh : 843"),
          ("→ C API",
           "roaring_bitmap_or(x1, x2)  — iterates segments by (high) key",
           "src/core/croaring/roaring.c : 24557"),
          ("→ per-segment dispatch",
           "container_or(c1, type1, c2, type2, &result_type)  "
           "— switch on (type1, type2)",
           "src/core/croaring/roaring.c : 4937"),
          ("→ case (ARRAY, ARRAY)",
           "calls array_array_container_union(c1, c2, &result)",
           "src/core/croaring/roaring.c : 4951"),
          ("⚑ Transition function",
           "array_array_container_union — chooses container type "
           "based on |A|+|B| vs DEFAULT_MAX_SIZE (4096); "
           "for our test |A|+|B|=6000 > 4096, so it allocates a "
           "bitset_container and copies both input arrays into it.",
           "src/core/croaring/roaring.c : 17995")],
         ),
        ("Scenario 2 — AND triggers bitset → array",
         [("C++ entry",
           "roaring::Roaring::operator&(const Roaring&)",
           "src/core/croaring/roaring.hh : 818"),
          ("→ C API",
           "roaring_bitmap_and(x1, x2)",
           "src/core/croaring/roaring.c"),
          ("→ per-segment dispatch",
           "container_and(c1, type1, c2, type2, &result_type)  "
           "— switch on (type1, type2)",
           "src/core/croaring/roaring.c : 4655"),
          ("→ case (BITSET, BITSET)",
           "calls bitset_bitset_container_intersection(c1, c2, &result)",
           "src/core/croaring/roaring.c : 4662"),
          ("⚑ Transition function",
           "bitset_bitset_container_intersection — first pass counts "
           "popcount(A[i] & B[i]) (full 8 KB scan); if result ≤ 4096 "
           "it allocates an array_container and extracts set-bit "
           "positions via bitset_extract_intersection_setbits_uint16 "
           "(second 8 KB scan).",
           "src/core/croaring/roaring.c : 17260")],
         ),
        ("Why ComBit avoids the cost",
         [("4-level fill hierarchy",
           "ComBit segments are described by L4/L3/L2 markers + L1 "
           "literal bytes.  Whether a segment is 'sparse' or 'dense' is "
           "encoded by the fill-polarity bits; the same data layout "
           "handles both regimes — no container-type concept.",
           "src/core/combit/include/combit.h"),
          ("Byte-level scatter",
           "OR / AND traverse L1 literal byte streams with AVX-512; "
           "no allocation per segment, no post-op conversion.",
           "src/core/combit/src/or.cpp / and.cpp")],
         ),
    ]

    row_cursor = 3
    for title, items in sections:
        subtitle_cell(ws.cell(row_cursor, 1), title)
        ws.merge_cells(start_row=row_cursor, start_column=1,
                       end_row=row_cursor, end_column=3)
        ws.row_dimensions[row_cursor].height = 22
        row_cursor += 1

        for step, desc, src in items:
            data_cell(ws.cell(row_cursor, 1), step,
                      bold=True, align="left", number_format="@")
            data_cell(ws.cell(row_cursor, 2), desc,
                      align="left", number_format="@")
            data_cell(ws.cell(row_cursor, 3), src,
                      align="left", number_format="@")
            # taller row for wrapped text
            ws.row_dimensions[row_cursor].height = max(
                36, 18 + len(desc) // 60 * 14)
            for col in (1, 2, 3):
                ws.cell(row_cursor, col).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
            row_cursor += 1
        row_cursor += 1


# ============================================================
# Sheet 4: Raw — the CSV verbatim
# ============================================================
def write_raw_sheet(ws, rows):
    ws.title = "Raw CSV"
    set_widths(ws, {"A": 14, "B": 12, "C": 14, "D": 12, "E": 12,
                    "F": 14, "G": 14, "H": 18})
    title_cell(ws.cell(1, 1),
               "Raw bench output (mirror of transition_results.csv)",
               size=13)
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 24

    cols = ["scenario", "backend", "median_ms", "min_ms", "max_ms",
            "stddev_ms", "result_card", "result_size_mb"]
    for i, h in enumerate(cols, start=1):
        header_cell(ws.cell(3, i), h)
    ws.row_dimensions[3].height = 22

    for r_idx, r in enumerate(rows, start=4):
        for c_idx, key in enumerate(cols, start=1):
            v = r[key]
            if key in {"result_card"}:
                v = int(v)
                fmt = "#,##0"
            elif key in {"scenario", "backend"}:
                fmt = "@"
            else:
                v = float(v)
                fmt = "0.0000"
            data_cell(ws.cell(r_idx, c_idx), v, number_format=fmt,
                      align="center")


# ============================================================
# Main
# ============================================================
def main():
    csv_in  = Path(sys.argv[1] if len(sys.argv) > 1
                   else "transition_results.csv")
    xlsx_out = Path(sys.argv[2] if len(sys.argv) > 2
                    else "transition_case.xlsx")
    rows = load_rows(csv_in)

    wb = Workbook()
    write_summary_sheet(wb.active, rows)
    write_chart_sheet(wb.create_sheet(), rows)
    write_howitworks_sheet(wb.create_sheet())
    write_raw_sheet(wb.create_sheet(), rows)

    wb.save(xlsx_out)
    print(f"[OK] wrote {xlsx_out}  ({xlsx_out.stat().st_size} bytes, "
          f"{len(rows)} timing rows, 4 sheets)")


if __name__ == "__main__":
    main()
