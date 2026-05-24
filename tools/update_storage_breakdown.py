#!/usr/bin/env python3
"""update_storage_breakdown.py

Parses /tmp/storage_breakdown.txt (output of the augmented benchmark run)
and writes detailed memory cells into results_report.xlsx for the WAH /
EWAH / Concise / Bitset / Bitset_AVX512 columns of each density block.

Cell format examples (matching the original Excel style):
  WAH      "fill_words: 46,310,901 (176.66 MB) | literal_words: 103,277,638
            (393.97 MB) | total: 598,360,556 bytes (570.64 MB)"
  EWAH     "marker_words: 769 (0.01 MB) | dirty_words: 113,083,114 (862.75 MB)
            | total: 904,664,272 bytes (862.75 MB)"
  Concise  "fill_words: 34,618,653 (132.06 MB) | literal_words: 60,804,445
            (231.95 MB) | total: 381,692,792 bytes (364.01 MB)"
  Bitset   "per_bitmap: 12,500,000 bytes (11.92 MB) × N | total: bytes (X MB)"
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment


SECTION_RE = re.compile(r"=== (\w+) c=(\d+) ===")
WAH_RE = re.compile(
    r"\[WAH Storage\] fill_words=(\d+) \(([\d.e+-]+) MB\) \| "
    r"literal_words=(\d+) \(([\d.e+-]+) MB\) \| "
    r"total_bytes=(\d+) \(([\d.e+-]+) MB\)"
)
EWAH_RE = re.compile(
    r"\[EWAH Storage\] marker_words=(\d+) \(([\d.e+-]+) MB\) \| "
    r"dirty_words=(\d+) \(([\d.e+-]+) MB\) \| "
    r"total_bytes=(\d+) \(([\d.e+-]+) MB\)"
)
CONCISE_RE = re.compile(
    r"\[Concise Storage\] fill_words=(\d+) \(([\d.e+-]+) MB\) \| "
    r"literal_words=(\d+) \(([\d.e+-]+) MB\) \| "
    r"total_bytes=(\d+) \(([\d.e+-]+) MB\)"
)
BITSET_RE = re.compile(
    r"\[Bitset Storage\] bitmaps=(\d+) \| "
    r"per_bitmap=(\d+) bytes \(([\d.e+-]+) MB\) \| "
    r"total_bytes=(\d+) \(([\d.e+-]+) MB\)"
)

# Backend (CLI name) → Excel column index + display formatter.
BACKEND_TO_COL = {
    "bitset":        2,
    "bitset_avx512": 3,
    "wah":           4,
    "ewah":          5,
    "concise":       6,
}

DENSITY_RE = re.compile(r"Density:\s*([0-9.]+)\s*\|\s*Loaded:\s*(\d+)")


def fmt_int(n):
    return f"{int(n):,}"


def fmt_mb(s):
    return f"{float(s):,.2f} MB"


def parse(path: Path):
    """Return dict: (backend, card) → formatted memory text per backend."""
    out: dict[tuple[str, int], str] = {}
    current_backend = None
    current_card = None
    for line in path.read_text().splitlines():
        m = SECTION_RE.match(line)
        if m:
            current_backend = m.group(1)
            current_card = int(m.group(2))
            continue
        if current_backend is None:
            continue

        if current_backend == "wah":
            m = WAH_RE.search(line)
            if m:
                fw, fmb, lw, lmb, tb, tmb = m.groups()
                out[("wah", current_card)] = (
                    f"fill_words: {fmt_int(fw)} ({fmt_mb(fmb)})\n"
                    f"literal_words: {fmt_int(lw)} ({fmt_mb(lmb)})\n"
                    f"total: {fmt_int(tb)} bytes ({fmt_mb(tmb)})"
                )
        elif current_backend == "ewah":
            m = EWAH_RE.search(line)
            if m:
                mw, mmb, dw, dmb, tb, tmb = m.groups()
                out[("ewah", current_card)] = (
                    f"marker_words: {fmt_int(mw)} ({fmt_mb(mmb)})\n"
                    f"dirty_words: {fmt_int(dw)} ({fmt_mb(dmb)})\n"
                    f"total: {fmt_int(tb)} bytes ({fmt_mb(tmb)})"
                )
        elif current_backend == "concise":
            m = CONCISE_RE.search(line)
            if m:
                fw, fmb, lw, lmb, tb, tmb = m.groups()
                out[("concise", current_card)] = (
                    f"fill_words: {fmt_int(fw)} ({fmt_mb(fmb)})\n"
                    f"literal_words: {fmt_int(lw)} ({fmt_mb(lmb)})\n"
                    f"total: {fmt_int(tb)} bytes ({fmt_mb(tmb)})"
                )
        elif current_backend in ("bitset", "bitset_avx512"):
            m = BITSET_RE.search(line)
            if m:
                nb, per_b, per_mb, tb, tmb = m.groups()
                out[(current_backend, current_card)] = (
                    f"bitmaps: {fmt_int(nb)}\n"
                    f"per_bitmap: {fmt_int(per_b)} bytes ({fmt_mb(per_mb)})\n"
                    f"total: {fmt_int(tb)} bytes ({fmt_mb(tmb)})"
                )
    return out


def main():
    storage_in = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/storage_breakdown.txt")
    xlsx_io    = Path(sys.argv[2] if len(sys.argv) > 2 else "results_report.xlsx")

    mem = parse(storage_in)
    print(f"[load] parsed {len(mem)} (backend, card) breakdowns")

    wb = openpyxl.load_workbook(xlsx_io)
    ws = wb.active

    written = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        m = DENSITY_RE.search(str(v))
        if not m:
            continue
        card = int(m.group(2))
        for be, col in BACKEND_TO_COL.items():
            text = mem.get((be, card))
            if not text:
                continue
            cell = ws.cell(r, col)
            cell.value = text
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
            written += 1

    # Make those rows tall enough to show 3 lines
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and DENSITY_RE.search(str(v)):
            ws.row_dimensions[r].height = 52

    print(f"[fill] {written} memory cells updated")
    wb.save(xlsx_io)
    print(f"[ok]   saved {xlsx_io}")


if __name__ == "__main__":
    main()
